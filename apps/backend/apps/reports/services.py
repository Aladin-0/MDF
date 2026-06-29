import csv
import io
from datetime import datetime, date, timedelta
from decimal import Decimal
from typing import Dict, Any, List

from django.db.models import Sum, Q, F, Case, When, Value, DecimalField, OuterRef, Subquery, IntegerField, FloatField
from django.db.models.functions import Coalesce, ExtractDay, Cast
from django.utils import timezone
from django.http import HttpResponse, StreamingHttpResponse
from django.template.loader import render_to_string

from openpyxl.styles import Font, PatternFill

from apps.inventory.models import Batch, StockLedger
from apps.inventory.utils import quantity_to_pack_display
from apps.purchases.models import PurchaseItem


class BatchWiseReportService:
    @staticmethod
    def get_batch_wise_report(
        outlet_id: str,
        report_type: str,
        date_from: date = None,
        date_to: date = None,
        search: str = None,
        product_id: str = None,
        supplier_id: str = None, # If distributor filtering is needed
        expiry_within_days: int = 90
    ) -> Dict[str, Any]:
        """
        Generate batch-wise tracking report data.
        report_types: current_stock, movement, near_expiry, expired, zero_stock
        """
        today = timezone.now().date()
        
        # Base QuerySet: Batches for the outlet
        qs = Batch.objects.filter(outlet_id=outlet_id).select_related('product')
        
        if product_id:
            qs = qs.filter(product_id=product_id)
            
        if search:
            qs = qs.filter(Q(product__name__icontains=search) | Q(batch_no__icontains=search))
            
        # We need to compute fractional total stock for filtering/sorting
        qs = qs.annotate(
            current_fractional_qty=Cast(F('qty_strips'), FloatField()) + Cast(F('qty_loose'), FloatField()) / Cast(Coalesce(F('pack_size'), 1), FloatField())
        )
            
        # Apply report_type specific filters early
        if report_type == 'near_expiry':
            target_date = today + timedelta(days=expiry_within_days)
            qs = qs.filter(expiry_date__gt=today, expiry_date__lte=target_date, current_fractional_qty__gt=0)
        elif report_type == 'expired':
            qs = qs.filter(expiry_date__lte=today, current_fractional_qty__gt=0)
        elif report_type == 'zero_stock':
            qs = qs.filter(current_fractional_qty__lte=0)
        elif report_type == 'current_stock':
            # Generally, current stock only shows >0 unless specified otherwise
            qs = qs.filter(current_fractional_qty__gt=0)
            
        # Supplier annotation
        supplier_sq = PurchaseItem.objects.filter(
            batch=OuterRef('pk')
        ).values('invoice__distributor__name')[:1]
        qs = qs.annotate(supplier_name_annotated=Subquery(supplier_sq))

        # Let's annotate movements if it's movement report
        if report_type == 'movement':
            # Base filters for stock ledger
            ledger_filter_base = Q()
            if date_from:
                ledger_filter_base &= Q(stock_ledger_entries__txn_date__gte=date_from)
            if date_to:
                ledger_filter_base &= Q(stock_ledger_entries__txn_date__lte=date_to)

            # Opening = Sum(qty_in) - Sum(qty_out) before date_from
            if date_from:
                qs = qs.annotate(
                    raw_opening_in=Coalesce(Sum('stock_ledger_entries__qty_in', filter=Q(stock_ledger_entries__txn_date__lt=date_from)), Decimal('0')),
                    raw_opening_out=Coalesce(Sum('stock_ledger_entries__qty_out', filter=Q(stock_ledger_entries__txn_date__lt=date_from)), Decimal('0')),
                ).annotate(
                    total_ledger_opening=F('raw_opening_in') - F('raw_opening_out')
                ).annotate(
                    opening_qty_raw=Case(
                        When(
                            raw_opening_in=Decimal('0'), raw_opening_out=Decimal('0'), created_at__date__lt=date_from,
                            then=Coalesce(F('opening_qty'), Decimal('0'))
                        ),
                        default=F('total_ledger_opening'),
                        output_field=DecimalField()
                    )
                )
            else:
                qs = qs.annotate(opening_qty_raw=Value(Decimal('0'), output_field=DecimalField()))
                
            # Range aggregations
            def get_sum(field, txn_type):
                return Coalesce(
                    Sum(f'stock_ledger_entries__{field}', filter=ledger_filter_base & Q(stock_ledger_entries__txn_type=txn_type)),
                    Decimal('0')
                )
                
            qs = qs.annotate(
                purchased_qty_raw=get_sum('qty_in', 'PURCHASE_IN'),
                sold_qty_raw=get_sum('qty_out', 'SALE_OUT'),
                sales_return_qty_raw=get_sum('qty_in', 'SALE_RETURN'),
                purchase_return_qty_raw=get_sum('qty_out', 'PURCHASE_RETURN'),
                adjustment_in_qty_raw=get_sum('qty_in', 'ADJUSTMENT_IN'),
                adjustment_out_qty_raw=get_sum('qty_out', 'ADJUSTMENT_OUT'),
            )
            
            qs = qs.annotate(
                closing_qty_raw=F('opening_qty_raw') + F('purchased_qty_raw') + F('sales_return_qty_raw') + F('adjustment_in_qty_raw') - F('sold_qty_raw') - F('purchase_return_qty_raw') - F('adjustment_out_qty_raw')
            )
        else:
            # For non-movement, raw_qty is just current qty
            qs = qs.annotate(
                opening_qty_raw=F('current_fractional_qty'),
                purchased_qty_raw=Value(Decimal('0'), output_field=DecimalField()),
                sold_qty_raw=Value(Decimal('0'), output_field=DecimalField()),
                sales_return_qty_raw=Value(Decimal('0'), output_field=DecimalField()),
                purchase_return_qty_raw=Value(Decimal('0'), output_field=DecimalField()),
                adjustment_in_qty_raw=Value(Decimal('0'), output_field=DecimalField()),
                adjustment_out_qty_raw=Value(Decimal('0'), output_field=DecimalField()),
                closing_qty_raw=F('current_fractional_qty')
            )

        # Ordering
        qs = qs.order_by('product__name', 'expiry_date')

        # Execute query
        batches = list(qs)
        
        # Format results
        rows = []
        warnings = []
        total_batches = len(batches)
        total_active_batches = 0
        total_near_expiry_batches = 0
        total_expired_batches = 0
        total_zero_stock_batches = 0
        total_closing_qty_raw = Decimal('0')
        total_stock_value = Decimal('0')
        
        for b in batches:
            days_to_expiry = (b.expiry_date - today).days if b.expiry_date else None
            
            if days_to_expiry is not None and days_to_expiry < 0:
                expiry_status = 'EXPIRED'
                total_expired_batches += 1
            elif days_to_expiry is not None and days_to_expiry <= 90:
                expiry_status = 'NEAR_EXPIRY'
                total_near_expiry_batches += 1
            else:
                expiry_status = 'ACTIVE'
                total_active_batches += 1
                
            if b.closing_qty_raw <= 0:
                total_zero_stock_batches += 1
                
            total_closing_qty_raw += Decimal(str(b.closing_qty_raw))
            
            stock_value = float(b.closing_qty_raw) * float(b.purchase_rate)
            total_stock_value += Decimal(str(stock_value))
            
            margin_pct = ((float(b.sale_rate) - float(b.purchase_rate)) / float(b.purchase_rate) * 100) if float(b.purchase_rate) > 0 else 0
            
            # Pack conversion helpers
            ps = b.pack_size or 1
            
            # Warnings
            if b.closing_qty_raw < 0:
                warnings.append({
                    "batch_id": str(b.id),
                    "batch_no": b.batch_no,
                    "medicine": b.product.name if b.product else "Custom Product",
                    "warning_type": "NEGATIVE_CLOSING_STOCK",
                    "message": f"Closing qty is negative ({float(b.closing_qty_raw)}) — possible missing opening entry"
                })
            if b.purchase_rate > b.mrp:
                warnings.append({
                    "batch_id": str(b.id),
                    "batch_no": b.batch_no,
                    "medicine": b.product.name if b.product else "Custom Product",
                    "warning_type": "PURCHASE_RATE_EXCEEDS_MRP",
                    "message": f"Purchase rate ₹{b.purchase_rate} exceeds MRP ₹{b.mrp}"
                })
            if b.sale_rate < b.purchase_rate:
                warnings.append({
                    "batch_id": str(b.id),
                    "batch_no": b.batch_no,
                    "medicine": b.product.name if b.product else "Custom Product",
                    "warning_type": "SALE_RATE_ABNORMAL",
                    "message": f"Sale rate ₹{b.sale_rate} is lower than purchase rate ₹{b.purchase_rate}"
                })
            
            rows.append({
                'batch_id': str(b.id),
                'product_id': str(b.product_id) if b.product_id else None,
                'medicine_name': b.product.name if b.product else "Custom Product",
                'batch_no': b.batch_no,
                'manufacturer_name': b.product.manufacturer if b.product else None,
                'supplier_name': getattr(b, 'supplier_name_annotated', None) or '',
                'mfg_date': b.mfg_date.isoformat() if b.mfg_date else None,
                'expiry_date': b.expiry_date.isoformat() if b.expiry_date else None,
                'days_to_expiry': days_to_expiry,
                'expiry_status': expiry_status,
                'mrp': float(b.mrp),
                'purchase_rate': float(b.purchase_rate),
                'sale_rate': float(b.sale_rate),
                'stock_value': stock_value,
                'margin_pct': margin_pct,
                'pack_size': ps,
                'pack_unit': b.pack_unit,
                'pack_type': b.pack_type,
                'rack_location': b.rack_location,
                
                'opening_qty_raw': float(b.opening_qty_raw),
                'purchased_qty_raw': float(b.purchased_qty_raw),
                'sold_qty_raw': float(b.sold_qty_raw),
                'sales_return_qty_raw': float(b.sales_return_qty_raw),
                'purchase_return_qty_raw': float(b.purchase_return_qty_raw),
                'adjustment_in_qty_raw': float(b.adjustment_in_qty_raw),
                'adjustment_out_qty_raw': float(b.adjustment_out_qty_raw),
                'closing_qty_raw': float(b.closing_qty_raw),
                
                'opening_qty_display': quantity_to_pack_display(b.opening_qty_raw, ps),
                'purchased_qty_display': quantity_to_pack_display(b.purchased_qty_raw, ps),
                'sold_qty_display': quantity_to_pack_display(b.sold_qty_raw, ps),
                'sales_return_qty_display': quantity_to_pack_display(b.sales_return_qty_raw, ps),
                'purchase_return_qty_display': quantity_to_pack_display(b.purchase_return_qty_raw, ps),
                'adjustment_in_qty_display': quantity_to_pack_display(b.adjustment_in_qty_raw, ps),
                'adjustment_out_qty_display': quantity_to_pack_display(b.adjustment_out_qty_raw, ps),
                'closing_qty_display': quantity_to_pack_display(b.closing_qty_raw, ps),
            })
            
        return {
            'data': rows,
            'warnings': warnings,
            'summary': {
                'total_batches': total_batches,
                'total_active_batches': total_active_batches,
                'total_near_expiry_batches': total_near_expiry_batches,
                'total_expired_batches': total_expired_batches,
                'total_zero_stock_batches': total_zero_stock_batches,
                'total_closing_qty_raw': float(total_closing_qty_raw),
                'total_stock_value': float(total_stock_value),
                'report_generated_at': timezone.now().isoformat(),
            }
        }

    @staticmethod
    def export_csv(report_data: Dict[str, Any]) -> HttpResponse:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="batch-wise-report-{timezone.now().strftime("%Y-%m-%d")}.csv"'
        response.write(u'\ufeff'.encode('utf8')) # BOM for Excel
        
        writer = csv.writer(response)
        writer.writerow([
            'Supplier', 'Medicine', 'Batch No', 'MFG Date', 'Expiry Date', 'Days to Expiry', 'Status',
            'MRP', 'Purchase Rate', 'Sale Rate', 'Margin %', 'Stock Value',
            'Opening', 'Purchased', 'Sold', 'Sales Return', 'Purchase Return', 'Adjustments', 'Closing'
        ])
        
        for row in report_data['data']:
            # Adjustments total = in - out
            adj = row['adjustment_in_qty_raw'] - row['adjustment_out_qty_raw']
            adj_display = quantity_to_pack_display(adj, row['pack_size'])['text']
            
            writer.writerow([
                row.get('supplier_name', ''),
                row['medicine_name'],
                row['batch_no'],
                row['mfg_date'] or '',
                row['expiry_date'] or '',
                row.get('days_to_expiry', ''),
                row['expiry_status'],
                row['mrp'],
                row['purchase_rate'],
                row['sale_rate'],
                round(row.get('margin_pct', 0), 2),
                round(row.get('stock_value', 0), 2),
                row['opening_qty_display']['text'],
                row['purchased_qty_display']['text'],
                row['sold_qty_display']['text'],
                row['sales_return_qty_display']['text'],
                row['purchase_return_qty_display']['text'],
                adj_display,
                row['closing_qty_display']['text'],
            ])
            
        return response

    @staticmethod
    def export_xlsx(report_data: Dict[str, Any], context: Dict[str, Any] = None) -> HttpResponse:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Batch Report"
        
        # Add metadata header
        if context:
            ws.append([f"Report Title: Batch Wise Tracking Report"])
            ws.append([f"Outlet: {context.get('outlet').name if context.get('outlet') else ''}"])
            ws.append([f"Date Range: {context.get('date_from', 'Start')} to {context.get('date_to', 'End')}"])
            ws.append([f"Report Type: {context.get('report_type', '')}"])
            ws.append([f"Generated By: {context.get('generated_by', '')}"])
            ws.append([f"Generated At: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"])
            ws.append([""])
        else:
            ws.append(["Report Title: Batch Wise Tracking Report"])
            ws.append(["Generated At", timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
            ws.append([""])
            ws.append([""])
            ws.append([""])
            ws.append([""])
            ws.append([""])
        
        headers = [
            'Supplier', 'Medicine', 'Batch No', 'MFG Date', 'Expiry Date', 'Days to Expiry', 'Status',
            'MRP', 'Purchase Rate', 'Sale Rate', 'Margin %', 'Stock Value',
            'Opening Strips', 'Opening Loose', 'Purchased Strips', 'Purchased Loose',
            'Sold Strips', 'Sold Loose', 'Sales Return Strips', 'Sales Return Loose',
            'Purchase Return Strips', 'Purchase Return Loose', 'Adjustments Strips', 'Adjustments Loose',
            'Closing Strips', 'Closing Loose'
        ]
        ws.append(headers)
        
        header_row_idx = ws.max_row
        
        # Style header
        header_font = Font(bold=True)
        for cell in ws[header_row_idx]:
            cell.font = header_font
            
        ws.freeze_panes = f'A{header_row_idx + 1}'
        
        status_fills = {
            'EXPIRED': PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid'),
            'NEAR_EXPIRY': PatternFill(start_color='FFE5CC', end_color='FFE5CC', fill_type='solid'),
            'ACTIVE': PatternFill(start_color='E5FFCC', end_color='E5FFCC', fill_type='solid')
        }
        
        for r_idx, row in enumerate(report_data['data'], start=header_row_idx + 1):
            adj = row['adjustment_in_qty_raw'] - row['adjustment_out_qty_raw']
            
            # Helper to split qty into strips and loose
            def split_qty(qty_raw, pack_size):
                import math
                qty_raw = float(qty_raw)
                strips = int(qty_raw) if qty_raw > 0 else math.ceil(qty_raw) if qty_raw < 0 else 0
                loose = round((qty_raw - strips) * pack_size)
                return strips, loose

            op_s, op_l = split_qty(row['opening_qty_raw'], row['pack_size'])
            pu_s, pu_l = split_qty(row['purchased_qty_raw'], row['pack_size'])
            so_s, so_l = split_qty(row['sold_qty_raw'], row['pack_size'])
            sr_s, sr_l = split_qty(row['sales_return_qty_raw'], row['pack_size'])
            pr_s, pr_l = split_qty(row['purchase_return_qty_raw'], row['pack_size'])
            ad_s, ad_l = split_qty(adj, row['pack_size'])
            cl_s, cl_l = split_qty(row['closing_qty_raw'], row['pack_size'])
            
            ws.append([
                row.get('supplier_name', ''),
                row['medicine_name'],
                row['batch_no'],
                row['mfg_date'] or '',
                row['expiry_date'] or '',
                row.get('days_to_expiry', ''),
                row['expiry_status'],
                row['mrp'],
                row['purchase_rate'],
                row['sale_rate'],
                round(row.get('margin_pct', 0), 2),
                round(row.get('stock_value', 0), 2),
                op_s, op_l, pu_s, pu_l, so_s, so_l,
                sr_s, sr_l, pr_s, pr_l, ad_s, ad_l,
                cl_s, cl_l
            ])
            
            # Color whole row based on status
            if row['expiry_status'] in status_fills:
                fill = status_fills[row['expiry_status']]
                for cell in ws[r_idx]:
                    cell.fill = fill
                
        # Append summary
        summary = report_data['summary']
        ws.append([])
        ws.append(["Total Batches:", summary['total_batches']])
        ws.append(["Active Batches:", summary['total_active_batches']])
        ws.append(["Near Expiry Batches:", summary['total_near_expiry_batches']])
        ws.append(["Expired Batches:", summary['total_expired_batches']])
        ws.append(["Zero Stock Batches:", summary['total_zero_stock_batches']])
        ws.append(["Total Inventory Value:", summary['total_stock_value']])
                
        # Auto-width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="batch-wise-report-{timezone.now().strftime("%Y-%m-%d")}.xlsx"'
        wb.save(response)
        return response

    @staticmethod
    def export_pdf(report_data: Dict[str, Any], context: Dict[str, Any]) -> HttpResponse:
        from weasyprint import HTML
        
        html_string = render_to_string('reports/batch_report.html', {
            'report_data': report_data,
            **context
        })
        
        pdf_file = HTML(string=html_string).write_pdf()
        
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="batch-wise-report-{timezone.now().strftime("%Y-%m-%d")}.pdf"'
        return response
