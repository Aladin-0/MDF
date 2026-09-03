import io
import hashlib
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import NotFound, PermissionDenied

from apps.core.models import Outlet
from apps.reports.models import GSTExportAudit, ITCReconciliationRun, ITCReconciliationResult

class ReconciliationExcelExportView(APIView):
    # Requires authentication and permission
    permission_classes = [IsAuthenticated]

    def get_current_outlet(self, request):
        return getattr(request.user, 'outlet', None)

    def get(self, request, fp):
        outlet = self.get_current_outlet(request)
        if not outlet:
            raise NotFound(detail="No outlet found")
            
        is_admin_or_super = getattr(request.user, 'role', '') in ('admin', 'super_admin')
        can_export = getattr(request.user, 'can_export_gst', False)
        if not (is_admin_or_super or can_export):
            raise PermissionDenied(detail="Missing GST export permission")

        run = ITCReconciliationRun.objects.filter(outlet=outlet, period=fp).last()
        if not run:
            raise NotFound(detail="No reconciliation run found for this period")

        results = ITCReconciliationResult.objects.filter(run=run)
        
        wb = openpyxl.Workbook()
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        
        # Helper to setup a sheet
        def setup_sheet(ws, title, headers):
            ws.title = title
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
            return ws

        # 1. Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Reconciliation Summary"
        
        ws_summary.append(["MediFlow Reconciliation Audit Report", ""])
        ws_summary.append(["Legal Entity", outlet.name])
        ws_summary.append(["GSTIN", outlet.gstin])
        ws_summary.append(["Period", fp])
        ws_summary.append(["Generated At", datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        ws_summary.append([])
        
        ws_summary.append(["Status", "Count"])
        matched = results.filter(match_status='MATCHED').count()
        missing_2b = results.filter(match_status='MISSING_IN_2B').count()
        missing_pr = results.filter(match_status='MISSING_IN_PR').count()
        mismatched = results.filter(match_status='MISMATCHED').count()
        
        ws_summary.append(["Total Invoices Matched", matched])
        ws_summary.append(["Total Mismatched (Value / Tax discrepancies)", mismatched])
        ws_summary.append(["Total Missing in Portal (2B)", missing_2b])
        ws_summary.append(["Total Missing in Books (PR)", missing_pr])
        
        for row in ws_summary.iter_rows(min_row=7, max_row=10, min_col=1, max_col=1):
            for cell in row:
                cell.font = Font(bold=True)

        # 2. Detailed Comparison Sheet
        ws_details = wb.create_sheet(title="Detailed Comparison")
        headers = [
            "Supplier Name", "Supplier GSTIN", "Invoice No", "Invoice Date", 
            "PR Taxable Value", "2B Taxable Value", "Taxable Diff", 
            "PR ITC Total", "2B ITC Total", "ITC Diff", "Status"
        ]
        setup_sheet(ws_details, "Detailed Comparison", headers)

        for r in results:
            pr = r.purchase_snapshot
            g2b = r.gstr2b_record

            pr_taxable = 0.0
            pr_itc = 0.0

            if pr and pr.snapshot_json:
                rates = pr.snapshot_json.get('items_by_rate', {})
                for rt, vals in rates.items():
                    pr_taxable += float(vals.get('taxable_amount', 0))
                    pr_itc += float(vals.get('igst', 0)) + float(vals.get('cgst', 0)) + float(vals.get('sgst', 0)) + float(vals.get('cess', 0))

            g2b_taxable = 0.0
            g2b_itc = 0.0

            if g2b:
                g2b_taxable = float(g2b.taxable_value)
                g2b_itc = float(g2b.igst) + float(g2b.cgst) + float(g2b.sgst) + float(g2b.cess)

            supplier_name = ""
            supplier_gstin = ""
            invoice_no = ""
            invoice_date = ""

            if g2b:
                supplier_name = g2b.supplier_name or ""
                supplier_gstin = g2b.supplier_gstin or ""
                invoice_no = g2b.invoice_number or ""
                invoice_date = g2b.invoice_date.strftime('%Y-%m-%d') if g2b.invoice_date else ""
            elif pr:
                supplier_name = pr.snapshot_json.get('customer_name', '')
                if not supplier_name:
                    supplier_name = pr.snapshot_json.get('supplier_name', '')
                supplier_gstin = pr.snapshot_json.get('supplier_gstin', '') or pr.gstin
                invoice_no = pr.document_number
                invoice_date = pr.document_date.strftime('%Y-%m-%d') if pr.document_date else ""

            status = r.match_status
            if status == 'MISSING_IN_2B':
                g2b_taxable = 0.0
                g2b_itc = 0.0
            elif status == 'MISSING_IN_PR':
                pr_taxable = 0.0
                pr_itc = 0.0
                
            taxable_diff = pr_taxable - g2b_taxable
            itc_diff = pr_itc - g2b_itc

            ws_details.append([
                supplier_name, supplier_gstin, invoice_no, invoice_date,
                pr_taxable, g2b_taxable, taxable_diff,
                pr_itc, g2b_itc, itc_diff, status
            ])

        # Save output
        out_stream = io.BytesIO()
        wb.save(out_stream)
        out_bytes = out_stream.getvalue()
        
        # Audit Log
        timestamp = datetime.now()
        output_file_hash = hashlib.sha256(out_bytes).hexdigest()
        
        GSTExportAudit.objects.create(
            actor=request.user if request.user.is_authenticated else None,
            outlet=outlet,
            period=fp,
            export_type='RECONCILIATION_EXCEL',
            output_file_hash=output_file_hash,
            validation_state={"message": "MediFlow Reconciliation Audit Generated"}
        )
        
        response = HttpResponse(out_bytes, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"MediFlow_Reconciliation_Audit_{outlet.gstin}_{fp}_{timestamp.strftime('%Y%m%d%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
